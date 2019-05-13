#### Padlock probe designer 2.0 by Kilian Zilkens####
# 2015-11-19

#TODO: output with Tm numbers
#TODO: read multiple sequences (fasta file)


#import mRNA sequence---------------------------

path="filename"

def start():
	global path
	print("This is a computational padlock probe design tool!")
	path=input("Please paste the full path to your mRNA.txt file (in FASTA format) here: ")
	Fasta=open(str(path),'r')
	mRNA=Fasta.read()
	seq=[]
	nucleobases = {'A':'A','a':'A','T':'T','t':'T','G':'G','g':'G','C':'C','c':'C','\n':'',' ':''} # clean up sequences, all in big letters, no space, no line break
	for item in mRNA:
		if item in nucleobases:
			seq.append(nucleobases[item])
		else:
			print("\nThe sequence contains an invalid character! Please try again!")    # NO FASTA HEADER
			start()
	clean_seq="".join(seq)
	start_chop(clean_seq)

##cDNA-chopping tool----------------------------

#global variables
P2=5
armlength=1
arm_tolerance=0
r=1

#constants
R=1.9872

#global variables for Tm-----------------------------------
C_oligo=5
C_salt=5
oligo_pos1=0
oligo_pos2=2
enthalpy_list=[]
entropy_list=[]
c=2
Tm_cor=1234
C_Mg=1
	
def start_chop(clean_seq):
	#get global variables
	global P2
	global armlength
	global arm_tolerance
	#length of probearms - only one arm, not both
	#armlength=eval(input("\nHow long should the arms of your padlock probes be?:"))
	armlength=20
	P2=armlength*2
	tolerance_fkt(clean_seq)
	
def tolerance_fkt(clean_seq):
	global arm_tolerance

	#input check
	tol_check=['Y','y','yes','N','n','no']
	#tolerance_YN=input("\nDo you want to add a tolerance range to the armlength? (Y/N):")
	tolerance_YN="Y"

    # correct for user input error
	if tolerance_YN in tol_check:
		pass
	else:
		print("\nPlease enter 'y' or 'n'!\n")
		tolerance_fkt(clean_seq)


	if tolerance_YN=='Y':
		#arm_tolerance=eval(input("How big should the tolerance be? (1-5):"))
		arm_tolerance=3
		check_tol(clean_seq)
	elif tolerance_YN=='y':
		#arm_tolerance=eval(input("How big should the tolerance be? (1-5):"))
		arm_tolerance=3
		check_tol(clean_seq)
	elif tolerance_YN=='yes':
		#arm_tolerance=eval(input("How big should the tolerance be? (1-5):"))
		arm_tolerance=3
		check_tol(clean_seq)

	if tolerance_YN=='N':
		chop(clean_seq)
	elif tolerance_YN=='n':
		chop(clean_seq)
	elif tolerance_YN=='no':
		chop(clean_seq)
		
#external arm_tolerance checking------------------------

def check_tol(clean_seq):
	global arm_tolerance
	if arm_tolerance>5:
		print("Please enter a number between 1 and 5!")
		tolerance_fkt(clean_seq)
	if arm_tolerance<0:
		print("Please enter a number between 1 and 5!")
		tolerance_fkt(clean_seq)
	if arm_tolerance==0:
		chop(clean_seq)
	else:
		chop_tol(clean_seq)



#---------------------chop without tolerance window----------------------
	
def chop(clean_seq):
	import openpyxl
	global path
	global P2
	global r
	P1=0
	print("\nGenerating output file... (may take a minute)\n")
	from openpyxl import Workbook
	from openpyxl import load_workbook
	wb = load_workbook('Padlocks.xlsx')
	ws=wb.create_sheet()
	#ws=wb.active(str(path))
	ws.title=str(path)[0:-4]
	while P2!= len(clean_seq)+1:
		AA=ws.cell(row=r,column=1)
		AA.value=clean_seq[P1:P2]
		ZZ=ws.cell(row=r+1,column=1)
		ZZ.value="endoflist"
		P1+=1
		P2+=1
		r+=1
	wb.save('Padlocks.xlsx')
	unregular_arms()

#chop with tolerance window----------------------
	
def chop_tol(clean_seq):
	import openpyxl
	global path
	global P2
	global arm_tolerance
	global r
	r=arm_tolerance+1
	P1=0
	print("\nGenerating output file... (may take a minute)\n")
	from openpyxl import workbook
	from openpyxl import load_workbook
	wb = load_workbook('Padlocks.xlsx')
	ws=wb.create_sheet()
	#ws=wb.active(str(path))
	ws.title=str(path)[0:-4]
	while P2!= len(clean_seq)+1:
		if arm_tolerance==1:
			AA=ws.cell(row=r-1,column=1)
			AA=clean_seq[P1:P2-1]
			BB=ws.cell(row=r,column=1)
			BB=clean_seq[P1:P2]
			CC=ws.cell(row=r+1,column=1)
			CC=clean_seq[P1:P2+1]
			ZZ=ws.cell(row=r+2,column=1)
			ZZ.value="endoflist"
			r+=3
		if arm_tolerance==2:
			AA=ws.cell(row=r-2,column=1)
			AA.value=clean_seq[P1:P2-2]
			BB=ws.cell(row=r-1,column=1)
			BB.value=clean_seq[P1:P2+-1]
			CC=ws.cell(row=r,column=1)
			CC.value=clean_seq[P1:P2]
			DD=ws.cell(row=r+1,column=1)
			DD.value=clean_seq[P1:P2+1]
			EE=ws.cell(row=r+2,column=1)
			EE.value=clean_seq[P1:P2+2]
			ZZ=ws.cell(row=r+3,column=1)
			ZZ.value="endoflist"
			r+=5
		if arm_tolerance==3:
			AA=ws.cell(row=r-3,column=1)
			AA.value=clean_seq[P1:P2-3]
			BB=ws.cell(row=r-2,column=1)
			BB.value=clean_seq[P1:P2-2]
			CC=ws.cell(row=r-1,column=1)
			CC.value=clean_seq[P1:P2-1]
			DD=ws.cell(row=r,column=1)
			DD.value=clean_seq[P1:P2]
			EE=ws.cell(row=r+1,column=1)
			EE.value=clean_seq[P1:P2+1]
			FF=ws.cell(row=r+2,column=1)
			FF.value=clean_seq[P1:P2+2]
			GG=ws.cell(row=r+3,column=1)
			GG.value=clean_seq[P1:P2+3]
			ZZ=ws.cell(row=r+4,column=1)
			ZZ.value="endoflist"
			r+=7
		if arm_tolerance==4:
			AA=ws.cell(row=r-4,column=1)
			AA.value=clean_seq[P1:P2-4]
			BB=ws.cell(row=r-3,column=1)
			BB.value=clean_seq[P1:P2-3]
			CC=ws.cell(row=r-2,column=1)
			CC.value=clean_seq[P1:P2-2]
			DD=ws.cell(row=r-1,column=1)
			DD.value=clean_seq[P1:P2-1]
			EE=ws.cell(row=r,column=1)
			EE.value=clean_seq[P1:P2]
			FF=ws.cell(row=r+1,column=1)
			FF.value=clean_seq[P1:P2+1]
			GG=ws.cell(row=r+2,column=1)
			GG.value=clean_seq[P1:P2+2]
			HH=ws.cell(row=r+3,column=1)
			HH.value=clean_seq[P1:P2+3]
			II=ws.cell(row=r+4,column=1)
			II.value=clean_seq[P1:P2+4]
			ZZ=ws.cell(row=r+5,column=1)
			ZZ.value="endoflist"
			r+=9
		if arm_tolerance==5:
			AA=ws.cell(row=r-5,column=1)
			AA.value=clean_seq[P1:P2-5]
			BB=ws.cell(row=r-4,column=1)
			BB.value=clean_seq[P1:P2-4]
			CC=ws.cell(row=r-3,column=1)
			CC.value=clean_seq[P1:P2-3]
			DD=ws.cell(row=r-2,column=1)
			DD.value=clean_seq[P1:P2-2]
			EE=ws.cell(row=r-1,column=1)
			EE.value=clean_seq[P1:P2-1]
			FF=ws.cell(row=r,column=1)
			FF.value=clean_seq[P1:P2]
			GG=ws.cell(row=r+1,column=1)
			GG.value=clean_seq[P1:P2+1]
			HH=ws.cell(row=r+2,column=1)
			HH.value=clean_seq[P1:P2+2]
			II=ws.cell(row=r+3,column=1)
			II.value=clean_seq[P1:P2+3]
			JJ=ws.cell(row=r+4,column=1)
			JJ.value=clean_seq[P1:P2+4]
			KK=ws.cell(row=r+5,column=1)
			KK.value=clean_seq[P1:P2+5]
			ZZ=ws.cell(row=r+6,column=1)
			ZZ.value="endoflist"
			r+=11
		P1+=1
		P2+=1
	wb.save('Padlocks.xlsx')
	unregular_arms()
	
def unregular_arms():
	from openpyxl import workbook
	from openpyxl import load_workbook
	wb = load_workbook('Padlocks.xlsx')
	ws=wb[str(path)[0:-4]]
	r=1
	check_content="OK_GO"
	while check_content != "endoflist":
		AA=ws.cell(row=r, column=1)
		full_arm=AA.value
		arm1=ws.cell(column=2,row=r)
		arm1.value=full_arm[0:len(full_arm)//2-2]
		ZZ1=ws.cell(column=2,row=r+1)
		ZZ1.value="endoflist"
		arm2=ws.cell(column=4,row=r)
		arm2.value=full_arm[len(full_arm)//2-2:len(full_arm)]
		ZZ2=ws.cell(column=4,row=r+1)
		ZZ2.value="endoflist"
		arm3=ws.cell(column=6,row=r)
		arm3.value=full_arm[0:len(full_arm)//2-1]
		ZZ3=ws.cell(column=6,row=r+1)
		ZZ3.value="endoflist"
		arm4=ws.cell(column=8,row=r)
		arm4.value=full_arm[len(full_arm)//2-1:len(full_arm)]
		ZZ4=ws.cell(column=8,row=r+1)
		ZZ4.value="endoflist"
		arm5=ws.cell(column=10,row=r)
		arm5.value=full_arm[0:len(full_arm)//2+1]
		ZZ5=ws.cell(column=10,row=r+1)
		ZZ5.value="endoflist"
		arm6=ws.cell(column=12,row=r)
		arm6.value=full_arm[len(full_arm)//2+1:len(full_arm)]
		ZZ6=ws.cell(column=12,row=r+1)
		ZZ6.value="endoflist"
		arm7=ws.cell(column=14,row=r)
		arm7.value=full_arm[0:len(full_arm)//2+2]
		ZZ7=ws.cell(column=14,row=r+1)
		ZZ7.value="endoflist"
		arm8=ws.cell(column=16,row=r)
		arm8.value=full_arm[len(full_arm)//2+2:len(full_arm)]
		ZZ8=ws.cell(column=16,row=r+1)
		ZZ8.value="endoflist"
		arm9=ws.cell(column=18,row=r)
		arm9.value=full_arm[0:len(full_arm)//2]
		ZZ9=ws.cell(column=18,row=r+1)
		ZZ9.value="endoflist"
		arm10=ws.cell(column=20,row=r)
		arm10.value=full_arm[len(full_arm)//2:len(full_arm)]
		ZZ10=ws.cell(column=20,row=r+1)
		ZZ10.value="endoflist"
		XX=ws.cell(column=1, row=r+1)
		check_content=XX.value
		r+=1
	wb.save('Padlocks.xlsx')
	start_tm()
	
#calculate melting temperature of an oligonucleotide-------------------------

def start_tm():
	global R
	global C_oligo
	global C_salt
	global oligo_pos1
	global oligo_pos2
	global enthalpy_list
	global entropy_list
	global r    # row
	global C_Mg
	global c    # column
	global Tm_cor
	r=1
	#C_oligo=eval(input("\nEnter the total concentration of your oligos in nM:"))/1000000000
	C_oligo=0.0000001
	print("\n>>>oligonucleotide concentration:" + str(C_oligo) + " M")
	#C_salt=eval(input("\nEnter your total concentration of monovalent cations in mM:"))/1000
	C_salt=0.075
	print("\n>>>ion concentration:" + str(C_salt) + " M")
	#C_Mg=eval(input("\nEnter your total concentration of Mg2+ ions in mM:"))/1000
	C_Mg=0.01
	print("\n>>>Mg2+ concentration:" + str(C_Mg) + " M")
	print("\nCalculating melting temperatures...\n")
	from openpyxl import workbook
	from openpyxl import load_workbook
	wb = load_workbook('Padlocks.xlsx')
	ws=wb[str(path)[0:-4]]
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	check_content="OK_GO"
	while check_content!="endoflist":
		AA=ws.cell(row=r, column=c)
		oligo=AA.value
		Tm(oligo)
		BB=ws.cell(row=r,column=c+1)
		BB.value=Tm_cor
		ZZ=ws.cell(row=r+1, column=c)
		check_content=ZZ.value
		r+=1
	c+=2
	r=1
	wb.save("Padlocks.xlsx")
	addlink()
	
def Tm(oligo):
	global R
	global C_oligo
	global C_salt
	global oligo_pos1
	global oligo_pos2
	global enthalpy_list
	global entropy_list
	global Tm_cor
	global C_Mg
	#calculate enthalpy
	#free enthalpies for base pairs
	H_AATT=-7.9
	H_ATTA=-7.2
	H_TAAT=-7.2
	H_CAGT=-8.5
	H_CTGA=-7.8
	H_GACT=-8.2
	H_GTCA=-8.4
	H_CGGC=-10.6
	H_GCCG=-9.8
	H_GGCC=-8.0
	AT_penalty_H=2.3
	GC_penalty_H=0.1
	#first arm:
	while oligo_pos2!=len(oligo)+1:
		if oligo[oligo_pos1:oligo_pos2]=="AA":
			enthalpy_list.append(H_AATT)
		if oligo[oligo_pos1:oligo_pos2]=="AT":
			enthalpy_list.append(H_ATTA)
		if oligo[oligo_pos1:oligo_pos2]=="AG":
			enthalpy_list.append(H_CTGA)
		if oligo[oligo_pos1:oligo_pos2]=="AC":
			enthalpy_list.append(H_GTCA)
		if oligo[oligo_pos1:oligo_pos2]=="TA":
			enthalpy_list.append(H_TAAT)
		if oligo[oligo_pos1:oligo_pos2]=="TC":
			enthalpy_list.append(H_GACT)
		if oligo[oligo_pos1:oligo_pos2]=="TG":
			enthalpy_list.append(H_CAGT)
		if oligo[oligo_pos1:oligo_pos2]=="TT":
			enthalpy_list.append(H_AATT)
		if oligo[oligo_pos1:oligo_pos2]=="GC":
			enthalpy_list.append(H_GCCG)
		if oligo[oligo_pos1:oligo_pos2]=="GA":
			enthalpy_list.append(H_GACT)
		if oligo[oligo_pos1:oligo_pos2]=="GT":
			enthalpy_list.append(H_GTCA)
		if oligo[oligo_pos1:oligo_pos2]=="GG":
			enthalpy_list.append(H_GGCC)
		if oligo[oligo_pos1:oligo_pos2]=="CG":
			enthalpy_list.append(H_CGGC)
		if oligo[oligo_pos1:oligo_pos2]=="CC":
			enthalpy_list.append(H_GGCC)
		if oligo[oligo_pos1:oligo_pos2]=="CA":
			enthalpy_list.append(H_CAGT)
		if oligo[oligo_pos1:oligo_pos2]=="CT":
			enthalpy_list.append(H_CTGA)
		oligo_pos1+=1
		oligo_pos2+=1
	if oligo[0]=="A":
		enthalpy_list.append(AT_penalty_H)
	if oligo[0]=="G":
		enthalpy_list.append(GC_penalty_H)	
	if oligo[0]=="T":
		enthalpy_list.append(AT_penalty_H)
	if oligo[0]=="C":
		enthalpy_list.append(GC_penalty_H)
	oligo_pos1=0
	oligo_pos2=2

#calculate entropy
	#free entropies for base pairs
	S_AATT=-22.2
	S_ATTA=-20.4
	S_TAAT=-21.3
	S_CAGT=-22.7
	S_CTGA=-21.0
	S_GACT=-22.2
	S_GTCA=-22.4
	S_CGGC=-27.2
	S_GCCG=-24.4
	S_GGCC=-19.9
	AT_penalty_S=4.1
	GC_penalty_S=-2.8
	while oligo_pos2!=len(oligo)+1:
		if oligo[oligo_pos1:oligo_pos2]=="AA":
			entropy_list.append(S_AATT)
		if oligo[oligo_pos1:oligo_pos2]=="AT":
			entropy_list.append(S_ATTA)
		if oligo[oligo_pos1:oligo_pos2]=="AG":
			entropy_list.append(S_CTGA)
		if oligo[oligo_pos1:oligo_pos2]=="AC":
			entropy_list.append(S_GTCA)
		if oligo[oligo_pos1:oligo_pos2]=="TA":
			entropy_list.append(S_TAAT)
		if oligo[oligo_pos1:oligo_pos2]=="TC":
			entropy_list.append(S_GACT)
		if oligo[oligo_pos1:oligo_pos2]=="TG":
			entropy_list.append(S_CAGT)
		if oligo[oligo_pos1:oligo_pos2]=="TT":
			entropy_list.append(S_AATT)
		if oligo[oligo_pos1:oligo_pos2]=="GC":
			entropy_list.append(S_GCCG)
		if oligo[oligo_pos1:oligo_pos2]=="GA":
			entropy_list.append(S_GACT)
		if oligo[oligo_pos1:oligo_pos2]=="GT":
			entropy_list.append(S_GTCA)
		if oligo[oligo_pos1:oligo_pos2]=="GG":
			entropy_list.append(S_GGCC)
		if oligo[oligo_pos1:oligo_pos2]=="CG":
			entropy_list.append(S_CGGC)
		if oligo[oligo_pos1:oligo_pos2]=="CC":
			entropy_list.append(S_GGCC)
		if oligo[oligo_pos1:oligo_pos2]=="CA":
			entropy_list.append(S_CAGT)
		if oligo[oligo_pos1:oligo_pos2]=="CT":
			entropy_list.append(S_CTGA)
		oligo_pos1+=1
		oligo_pos2+=1
	if oligo[0]=="A":
		entropy_list.append(AT_penalty_S)
	if oligo[0]=="G":
		entropy_list.append(GC_penalty_S)
	if oligo[0]=="T":
		entropy_list.append(AT_penalty_S)
	if oligo[0]=="C":
		entropy_list.append(GC_penalty_S)
#calculate Tm--------------------------------------------------------------------
#formula segment
	import math
	tot_enthalpy=math.fsum(enthalpy_list)
	tot_entropy=math.fsum(entropy_list)
	C_MVC=C_salt+(3.795*((C_Mg)**0.5))
	Tm=(tot_enthalpy*1000)/(tot_entropy + (R * math.log(C_oligo)))-273.15
	Tm_salt=Tm+(16.6*math.log10(C_MVC))
#correction for 20% formamide
	Tm_cor=Tm_salt-(0.72*20)
	#print(enthalpy_list)
	#print(entropy_list)
#reset the global lists and variables
	oligo_pos1=0
	oligo_pos2=2
	enthalpy_list=[]
	entropy_list=[]
	#print("\n")
	#print(len(enthalpy_list))
	#print(len(entropy_list))
	#print("\n")
	#print(C_oligo)
	#print(C_salt)
	#print("\n")
	#print(tot_enthalpy)
	#print(tot_entropy)
	#print(oligo)
	#print (C_MVC)
	#print("\n\n")
	#print(Tm)
	#print("corrected for 20% formamide:" + str(Tm_cor))
	
#construct your complete padlocks<-----------------------------------------------------------------------------------
def addlink():
	global path
	from openpyxl import workbook
	from openpyxl import load_workbook
	from openpyxl.cell import get_column_letter
	wb = load_workbook('Padlocks.xlsx')
	ws=wb[str(path)[0:-4]]
#temperature range parameters
	T_min=48
	T_max=52
#sequencers of linkers and anchorprimer binding site
	barcode=input("\nPlease enter the desired barcode for this probe: ")
	linker_first="TCCTCTATGATTACTGAC"
	linker_second="CTATCTTCTTT"
	ap_binding="TGCGTCTATTTAGTGGAGCC"
	#barcode=input("Which barcode should this padlock probe contain?": )
	linkadd=open("added_linkers.txt",'w')
	row=1
	c=3
	check_content="OK_GO"
	print("\nCreating full length probes...")
	while check_content!="endoflist":
		#define cells with Tm values
		AA=ws.cell(column=c, row=row)
		BB=ws.cell(column=c+2, row=row)
		CC=ws.cell(column=c+4, row=row)
		DD=ws.cell(column=c+6, row=row)
		EE=ws.cell(column=c+8, row=row)
		FF=ws.cell(column=c+10, row=row)
		GG=ws.cell(column=c+12, row=row)
		HH=ws.cell(column=c+14, row=row)
		II=ws.cell(column=c+16, row=row)
		JJ=ws.cell(column=c+18, row=row)
		#define cells with padlock arms
		A1=ws.cell(column=c-1,row=row)
		A2=ws.cell(column=c+1,row=row)
		A3=ws.cell(column=c+3,row=row)
		A4=ws.cell(column=c+5,row=row)
		A5=ws.cell(column=c+7,row=row)
		A6=ws.cell(column=c+9,row=row)
		A7=ws.cell(column=c+11,row=row)
		A8=ws.cell(column=c+13,row=row)
		A9=ws.cell(column=c+15,row=row)
		A10=ws.cell(column=c+17,row=row)
		arm1=A1.value
		arm2=A2.value
		arm3=A3.value
		arm4=A4.value
		arm5=A5.value
		arm6=A6.value
		arm7=A7.value
		arm8=A8.value
		arm9=A9.value
		arm10=A10.value
		#create full probes and move to .txt
		if T_min<AA.value<T_max and T_min<BB.value<T_max:
			quad_check=str(arm2) + str(linker_first) + str(ap_binding) + str(barcode) + str(linker_second) + str(arm1)
			if "AAAA" in quad_check:
				pass
			elif "TTTT" in quad_check:
				pass
			elif "CCCC" in quad_check:
				pass
			elif "GGGG" in quad_check:
				pass
			else:
				linkadd.write(str(quad_check) + "\n")
		if T_min<CC.value<T_max and T_min<DD.value<T_max:
			quad_check=(str(arm4) + str(linker_first) + str(ap_binding) + str(barcode) + str(linker_second) + str(arm3))
			if "AAAA" in quad_check:
				pass
			elif "TTTT" in quad_check:
				pass
			elif "CCCC" in quad_check:
				pass
			elif "GGGG" in quad_check:
				pass
			else:
				linkadd.write(str(quad_check) + "\n")
		if T_min<EE.value<T_max and T_min<FF.value<T_max:
			quad_check=(str(arm6) + str(linker_first) + str(ap_binding) + str(barcode) + str(linker_second) + str(arm5))
			if "AAAA" in quad_check:
				pass
			elif "TTTT" in quad_check:
				pass
			elif "CCCC" in quad_check:
				pass
			elif "GGGG" in quad_check:
				pass
			else:
				linkadd.write(str(quad_check) + "\n")
		if T_min<GG.value<T_max and T_min<HH.value<T_max:
			quad_check=(str(arm8) + str(linker_first) + str(ap_binding) + str(barcode) + str(linker_second) + str(arm7))
			if "AAAA" in quad_check:
				pass
			elif "TTTT" in quad_check:
				pass
			elif "CCCC" in quad_check:
				pass
			elif "GGGG" in quad_check:
				pass
			else:
				linkadd.write(str(quad_check) + "\n")
		if T_min<II.value<T_max and T_min<JJ.value<T_max:
			quad_check=(str(arm10) + str(linker_first) + str(ap_binding) + str(barcode) + str(linker_second) + str(arm9))
			if "AAAA" in quad_check:
				pass
			elif "TTTT" in quad_check:
				pass
			elif "CCCC" in quad_check:
				pass
			elif "GGGG" in quad_check:
				pass
			else:
				linkadd.write(str(quad_check) + "\n")
		ZZ=ws.cell(row=row+1, column=2)
		check_content=ZZ.value
		row+=1
	linkadd.flush()
	linkadd.close
	check_self()
	
def check_self():	
#-----------------------does user want complementarity check?--------------------------------------------------------------------
	comp_check=['Y','y','yes','N','n','no']
	#comp_YN=input("\nDo you want to check your probes for complementarity? (Y/N):")
	comp_YN="Y"
	if comp_YN in comp_check:
		pass
	else:
		print("\nPlease enter 'y' or 'n'!\n")
		check_self()

	if comp_YN=='N':
		thanks()
		pass
	elif comp_YN=='n':
		thanks()
		pass
	elif comp_YN=='no':
		thanks()
		pass
	else:
		#self_tol=eval(input("\nSet threshold size for complementary elements:"))
		self_tol=7


#------------------------check for complementarity---------------------------------------------------------------------------------
	print("Checking for self-complementarity...")
	infile=open("added_linkers.txt",'r')
	#empty=open("Final_probes.lis",'w')
	#empty.write("")
	#empty.flush()
	#empty.close()
	finprob=open(str(path)[0:-4]+".lis",'a')
	for line in infile:
		fullprobe=line
		comp_pen="false"
		A=0
		B=self_tol+1
		while B<(len(fullprobe)):   # consider escaping loop earlier
			armpiece=fullprobe[A:B]
			comp_arm=armpiece.translate({ord('A'):ord('T'),ord('T'):'A',ord('G'):'C',ord('C'):'G'}) # TODO: replace string
			if str(comp_arm) in fullprobe:#[B-1:-1] only rest of the probe --> should not be needed
				comp_pen="true"
			elif str(comp_arm)[::-1] in fullprobe:#[B-1:-1]
				comp_pen="true"
			A+=1
			B+=1
		if comp_pen=="false":
			finprob.write(str(fullprobe))
		else:
			pass
	print("\nOutput file generated.")
	thanks()
			
def thanks():
	print("\nThank you for using this padlock probe tool!\n")
	return 0
	
start()
#check_self()